import datetime as dt
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border

from core import cleaning
from config.settings import CFG

from core.dedupe_internal import dedupe_internal
from core.dedupe_external import dedupe_external

def clean_colname(name):
    return cleaning.clean_basic(str(name)).replace('\t', '').replace('\n', '').replace('\r', '').strip()

def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", FutureWarning)
        df = df.applymap(cleaning.clean_basic)
        df.columns = [clean_colname(c) for c in df.columns]
        return df

def extract_tables_multi(df: pd.DataFrame, logger=None):
    tables = []
    idx = 0
    nrows = df.shape[0]
    while idx < nrows:
        title_row = None
        for i in range(idx, nrows):
            row = df.iloc[i].astype(str).tolist()
            if any(cell.strip().startswith("【") for cell in row if isinstance(cell, str)):
                title_row = i
                break
        if title_row is None:
            break
        header_row = None
        for i in range(title_row + 1, nrows):
            row = df.iloc[i].astype(str).tolist()
            if any("No." in cell for cell in row):
                header_row = i
                break
        if header_row is None:
            idx = title_row + 1
            continue
        data_start = header_row + 1
        data_end = nrows
        for i in range(data_start, nrows):
            row = df.iloc[i].astype(str).tolist()
            if all(cell == "" or cell.lower() == "nan" for cell in row) or any(cell.strip().startswith("【") for cell in row):
                data_end = i
                break
        headers = [clean_colname(h) for h in df.iloc[header_row]]
        df_data = df.iloc[data_start:data_end].reset_index(drop=True)
        df_data.columns = headers
        df_data = df_data.ffill()
        title = next(cell.strip() for cell in df.iloc[title_row] if isinstance(cell, str) and cell.strip().startswith("【"))
        df_data["請求公演名"] = title
        must_have = ["氏名", "メールアドレス", "閲覧用URL"]
        if all(col in df_data.columns for col in must_have):
            df_data = clean_dataframe(df_data)
            for col in ["氏名", "メールアドレス"]:
                df_data = df_data[~df_data[col].isin([col, "", None, pd.NA])]
            df_data = df_data[
                (df_data["氏名"].astype(str).str.strip() != "") &
                (df_data["メールアドレス"].astype(str).str.strip() != "") &
                (df_data["閲覧用URL"].astype(str).str.strip() != "")
            ]
            tables.append(df_data)
        idx = data_end + 1
    if logger:
        logger.info(f"[extract_tables_multi] 抽出テーブル数: {len(tables)}")
        for i, t in enumerate(tables):
            logger.info(f"  テーブル{i+1}: {len(t)}件, カラム: {list(t.columns)}")
    return tables

def load_input_excel(path: Path, logger=None) -> pd.DataFrame:
    tables = []
    xls = pd.read_excel(path, sheet_name=None, header=None, dtype=str, engine="openpyxl")
    for sheet, df in xls.items():
        if logger:
            logger.info(f"[DEBUG] シート名: {sheet}")
        else:
            print(f"[DEBUG] シート名: {sheet}")
        ts = extract_tables_multi(df, logger)
        tables.extend(ts)
    df_all = pd.concat(tables, ignore_index=True) if tables else pd.DataFrame()
    df_all["_入力順"] = range(len(df_all))
    return df_all

def style_excel(path: Path, font_name: str):
    wb = load_workbook(path)
    ft = Font(name=font_name)
    nob = Border()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                c.font = ft
                c.border = nob
                c.number_format = "@"
    wb.save(path)

def append_and_save(df_new: pd.DataFrame, serial_gen, logger=None, overwrite=False):
    out_xlsx = Path(CFG["paths"]["output_excel"])
    bak_dir  = Path(CFG["paths"]["bak_dir"])
    bak_dir.mkdir(parents=True, exist_ok=True)
    ts = dt.datetime.now().strftime("%Y-%m-%d_%H%M")
    today_str = dt.datetime.now().strftime("%Y-%m-%d")
    if out_xlsx.exists():
        shutil.copy2(out_xlsx, bak_dir / f"bak_{out_xlsx.stem}_{ts}.xlsx")

    key_cols = ["氏名", "メールアドレス"]

    if "閲覧用URL" not in df_new.columns:
        df_new["閲覧用URL"] = ""
    if "_入力順" not in df_new.columns:
        df_new["_入力順"] = range(len(df_new))
    if "履歴" not in df_new.columns:
        df_new["履歴"] = ""

    df_person = dedupe_internal(df_new, logger=logger)

    if out_xlsx.exists() and not overwrite:
        try:
            df_existing = pd.read_excel(out_xlsx, engine="openpyxl", dtype=str)
            df_existing = clean_dataframe(df_existing).fillna("").astype(str)
            df_person = dedupe_external(df_person, df_existing, today_str, logger=logger)
        except Exception as e:
            if logger:
                logger.error(f"既存Excelの読み込みに失敗しました: {e}")

    for i, row in df_person.iterrows():
        if not row.get("管理番号") or str(row["管理番号"]).strip() == "":
            df_person.at[i, "管理番号"] = serial_gen.get_serial(row["氏名"], row["メールアドレス"])

    col_order = [
        "管理番号", "氏名", "メールアドレス", "電話番号", "郵便番号",
        "登録住所", "本人確認登録郵便番号", "本人確認登録時住所",
        "請求公演名", "件数", "備考", "履歴"
    ]
    for col in col_order:
        if col not in df_person.columns:
            df_person[col] = ""
    df_person = df_person[col_order]
    df_person = clean_dataframe(df_person).astype(str)

    tmp_xlsx = out_xlsx.with_name(out_xlsx.stem + "_tmp.xlsx")
    try:
        out_xlsx.parent.mkdir(parents=True, exist_ok=True)
        df_person.to_excel(tmp_xlsx, index=False)
        tmp_xlsx.replace(out_xlsx)
        style_excel(out_xlsx, CFG["excel"]["font_name"])
    except PermissionError:
        if logger:
            logger.error("CustList.xlsx を開いているため書き込めません。閉じてから再実行してください。")
        return

    if "管理番号" not in df_new.columns:
        df_new["管理番号"] = ""
    person_to_no = dict(zip(
        zip(df_person["氏名"], df_person["メールアドレス"], df_person["請求公演名"]),
        df_person["管理番号"]
    ))
    df_new["管理番号"] = df_new.apply(
        lambda row: person_to_no.get((row["氏名"], row["メールアドレス"], row["請求公演名"]), ""),
        axis=1
    )
    csv_col_order = [
        "管理番号", "氏名", "メールアドレス", "電話番号", "郵便番号",
        "登録住所", "本人確認登録郵便番号", "本人確認登録時住所",
        "請求公演名", "備考", "閲覧用URL"
    ]
    for col in csv_col_order:
        if col not in df_new.columns:
            df_new[col] = ""
    csv_df = df_new[csv_col_order]

    csv_name = CFG["paths"]["csv_pattern"].replace(
        "{yyyymmdd}", dt.datetime.today().strftime("%Y-%m-%d_%H%M")
    )
    Path(csv_name).parent.mkdir(parents=True, exist_ok=True)
    csv_df = clean_dataframe(csv_df).astype(str)
    tmp_csv_name = Path(csv_name).with_name(Path(csv_name).stem + "_tmp.csv")
    try:
        csv_df.to_csv(tmp_csv_name, index=False, encoding=CFG["csv"]["encoding"])
        tmp_csv_name.replace(csv_name)
    except Exception as e:
        if logger:
            logger.error(f"CSV一時ファイルの書き出しに失敗しました。元ファイルは壊れていません。: {e}")
        return

    if logger:
        logger.info(f"追記完了：{len(df_person)} 人 / {len(df_new)} URL")
        logger.info(f"Excel 保存: {out_xlsx}")
        logger.info(f"CSV 出力 : {csv_name}")

def load_person_map(logger=None) -> dict[str, str]:
    p = Path(CFG["paths"]["output_excel"])
    if not p.exists():
        return {}
    df = pd.read_excel(p, engine="openpyxl", dtype=str).fillna("").astype(str)
    df = clean_dataframe(df)
    df["key"] = df["氏名"].astype(str) + "|" + df["メールアドレス"].astype(str) + "|" + df["請求公演名"].astype(str)
    return dict(zip(df["key"], df["管理番号"].astype(str)))
