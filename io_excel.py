import datetime as dt
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border

import cleaning
from settings import CFG

def clean_colname(name):
    """カラム名もクリーニング（clean_basic + strip/tabs/newline除去）"""
    return cleaning.clean_basic(str(name)).replace('\t', '').replace('\n', '').replace('\r', '').strip()

def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", FutureWarning)
        df = df.applymap(cleaning.clean_basic)
        df.columns = [clean_colname(c) for c in df.columns]
        return df

def extract_tables_multi(df: pd.DataFrame, logger=None):
    # 複数テーブル抽出用
    tables = []
    idx = 0
    nrows = df.shape[0]
    while idx < nrows:
        # タイトル行（「【」で始まるものを想定）を探す
        title_row = None
        for i in range(idx, nrows):
            row = df.iloc[i].astype(str).tolist()
            if any(cell.strip().startswith("【") for cell in row if isinstance(cell, str)):
                title_row = i
                break
        if title_row is None:
            break  # 残りにタイトルがなければ終了
        # ヘッダー行（No.がある行）を探す
        header_row = None
        for i in range(title_row + 1, nrows):
            row = df.iloc[i].astype(str).tolist()
            if any("No." in cell for cell in row):
                header_row = i
                break
        if header_row is None:
            idx = title_row + 1
            continue  # 次を探す
        # データ部分（次の空行または次のタイトル行まで）
        data_start = header_row + 1
        data_end = nrows
        for i in range(data_start, nrows):
            row = df.iloc[i].astype(str).tolist()
            # 空行またはタイトル行でデータ終了
            if all(cell == "" or cell.lower() == "nan" for cell in row) or any(cell.strip().startswith("【") for cell in row):
                data_end = i
                break
        # カラム名クリーニング
        headers = [clean_colname(h) for h in df.iloc[header_row]]
        # データ部を抽出
        df_data = df.iloc[data_start:data_end].reset_index(drop=True)
        df_data.columns = headers
        # ffill（前方補完）で結合セルの値を全行に展開
        df_data = df_data.ffill()
        # 請求公演名を付与
        title = next(cell.strip() for cell in df.iloc[title_row] if isinstance(cell, str) and cell.strip().startswith("【"))
        df_data["請求公演名"] = title
        # 必須カラムチェック
        must_have = ["氏名", "メールアドレス", "閲覧用URL"]
        if all(col in df_data.columns for col in must_have):
            df_data = clean_dataframe(df_data)
            # 空欄行を除外
            for col in ["氏名", "メールアドレス"]:
                df_data = df_data[~df_data[col].isin([col, "", None, pd.NA])]
            df_data = df_data[
                (df_data["氏名"].astype(str).str.strip() != "") &
                (df_data["メールアドレス"].astype(str).str.strip() != "") &
                (df_data["閲覧用URL"].astype(str).str.strip() != "")
            ]
            tables.append(df_data)
        idx = data_end + 1  # 次のタイトル以降へ進む
    # ログ
    if logger:
        logger.info(f"[extract_tables_multi] 抽出テーブル数: {len(tables)}")
        for i, t in enumerate(tables):
            logger.info(f"  テーブル{i+1}: {len(t)}件, カラム: {list(t.columns)}")
    return tables

def load_input_excel(path: Path, logger=None) -> pd.DataFrame:
    tables = []
    xls = pd.read_excel(path, sheet_name=None, header=None, dtype=str, engine="openpyxl")
    for sheet, df in xls.items():
        if logger:
            logger.info(f"[DEBUG] シート名: {sheet}")
        else:
            print(f"[DEBUG] シート名: {sheet}")
        ts = extract_tables_multi(df, logger)
        tables.extend(ts)
    df_all = pd.concat(tables, ignore_index=True) if tables else pd.DataFrame()
    df_all["_入力順"] = range(len(df_all))
    return df_all

def style_excel(path: Path, font_name: str):
    wb = load_workbook(path)
    ft = Font(name=font_name)
    nob = Border()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                c.font = ft
                c.border = nob
                c.number_format = "@"
    wb.save(path)

def merge_person_records(df_person, df_existing, today_str, logger=None):
    """
    - 請求公演名は「部分一致」判定（区切り|でsplit）
    - 件数を加算＆公演名追記（部分一致無い場合）
    - 履歴（複数一致時は全管理番号）を記載して「新規追加」
    - 既存分への追記＆履歴付き新規分の追加を両立
    """
    added_records = []
    existing_map = {(row["氏名"], row["メールアドレス"]): i for i, row in df_existing.iterrows()}
    for idx, row in df_person.iterrows():
        key = (row["氏名"], row["メールアドレス"])
        new_titles = set(str(row["請求公演名"]).split("|")) if row["請求公演名"] else set()
        found_partial = False
        duplicate_ids = []
        if key in existing_map:
            # 同一人物の既存全レコード（請求公演名:部分一致）抽出
            exist_matches = df_existing[(df_existing["氏名"] == key[0]) & (df_existing["メールアドレス"] == key[1])]
            for _, exist_row in exist_matches.iterrows():
                exist_titles = set(str(exist_row["請求公演名"]).split("|")) if exist_row["請求公演名"] else set()
                if any(et in new_titles or nt in exist_titles for et in exist_titles for nt in new_titles):
                    duplicate_ids.append(str(exist_row["管理番号"]))
            if duplicate_ids:
                # 履歴つき新規追加
                hist = f"過去に同一人物・公演のレコード有り（管理番号:{'/'.join(sorted(set(duplicate_ids)))})"
                rec = row.copy()
                rec["履歴"] = hist
                added_records.append(rec)
                continue  # 件数カウント・追記せず新規
            else:
                # 部分一致無し → 件数加算＆公演名追記
                i = existing_map[key]
                exist_row = df_existing.loc[i]
                # 件数加算
                try:
                    new_cnt = int(exist_row["件数"]) + int(row.get("件数", 1))
                except Exception:
                    new_cnt = 1
                # 請求公演名追記
                all_titles = set(str(exist_row["請求公演名"]).split("|")) | new_titles
                merged_titles = "|".join(sorted(t for t in all_titles if t))
                prev_hist = str(exist_row.get("履歴") or "")
                upd_str = f"{today_str}:追記"
                upd_col = prev_hist + "|" + upd_str if prev_hist else upd_str
                for col in df_person.columns:
                    if col == "件数":
                        df_existing.at[i, col] = new_cnt
                    elif col == "請求公演名":
                        df_existing.at[i, col] = merged_titles
                    elif col == "履歴":
                        df_existing.at[i, col] = upd_col
                # 追加しない
        else:
            # 既存に同一人物無し→そのまま新規
            added_records.append(row.copy())
    # DataFrame返す
    df_add = pd.DataFrame(added_records)
    return df_add

def export_duplicate_records(df, base_dir="output", today_str=None):
    if df.empty:
        return
    if today_str is None:
        today_str = dt.datetime.today().strftime("%Y-%m-%d_%H%M")
    fname = f"CustList_{today_str}_removed.xlsx"
    out_path = Path(base_dir) / fname
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(out_path, index=False)

def remove_internal_duplicates(df_new):
    """
    同一ファイル（プログラム1回実行）内の重複判定:
    ・氏名＋メール＋請求公演名＋URLが一致→完全重複→無視
    ・氏名＋メール＋請求公演名が一致、URLが違う→件数カウント＆公演名追記
    """
    key_cols = ["氏名", "メールアドレス", "請求公演名"]
    url_col = "閲覧用URL"
    urlset_by_key = {}
    dedup_rows = []
    for _, row in df_new.iterrows():
        key = tuple(row[k] for k in key_cols)
        url = row[url_col]
        if key not in urlset_by_key:
            urlset_by_key[key] = set()
        if url in urlset_by_key[key]:
            continue  # 完全重複
        urlset_by_key[key].add(url)
        dedup_rows.append(row)
    df_person = pd.DataFrame(dedup_rows)
    return df_person

def append_and_save(df_new: pd.DataFrame, serial_gen, logger=None, overwrite=False):
    out_xlsx = Path(CFG["paths"]["output_excel"])
    bak_dir  = Path(CFG["paths"]["bak_dir"])
    bak_dir.mkdir(parents=True, exist_ok=True)
    ts = dt.datetime.now().strftime("%Y-%m-%d_%H%M")
    today_str = dt.datetime.now().strftime("%Y-%m-%d")
    if out_xlsx.exists():
        shutil.copy2(out_xlsx, bak_dir / f"bak_{out_xlsx.stem}_{ts}.xlsx")

    key_cols = ["氏名", "メールアドレス"]

    if "閲覧用URL" not in df_new.columns:
        df_new["閲覧用URL"] = ""
    if "_入力順" not in df_new.columns:
        df_new["_入力順"] = range(len(df_new))
    if "履歴" not in df_new.columns:
        df_new["履歴"] = ""

    # 1. 同一ファイル内の重複除去
    df_person = remove_internal_duplicates(df_new)
    # 原始データで人×URL件数
    if logger:
        logger.info(f"==== [原始データ:人×URL] レコード数: {len(df_person)} ====")
        logger.info(df_person[[*key_cols, "閲覧用URL"]].head(30))

    # 2. URL件数集計
    urlset_df = (
        df_person
        .groupby(key_cols)["閲覧用URL"]
        .apply(lambda x: set([v for v in x if pd.notnull(v) and str(v).strip() != ""]))
        .reset_index()
    )
    urlset_df["件数"] = urlset_df["閲覧用URL"].apply(len)

    # 3. 代表情報抽出
    agg_dict = {
        "管理番号": "first",
        "電話番号": "first",
        "郵便番号": "first",
        "登録住所": "first",
        "本人確認登録郵便番号": "first",
        "本人確認登録時住所": "first",
        "請求公演名": "first",
        "備考": "first",
        "_入力順": "first",
        "履歴": "first"
    }
    df_person2 = df_person.groupby(key_cols, as_index=False).agg(agg_dict)
    df_person2 = pd.merge(df_person2, urlset_df[[*key_cols, "件数"]], on=key_cols, how="left")
    df_person2["件数"] = df_person2["件数"].fillna(0).astype(int)
    df_person2 = df_person2.sort_values("_入力順").reset_index(drop=True)

    # 管理番号を採番（現行通り）
    for i, row in df_person2.iterrows():
        if not row.get("管理番号") or str(row["管理番号"]).strip() == "":
            df_person2.at[i, "管理番号"] = serial_gen.get_serial(row["氏名"], row["メールアドレス"])

    # 既存Excelとのマージ＆重複追記対応
    added_records = pd.DataFrame()
    if out_xlsx.exists() and not overwrite:
        try:
            df_existing = pd.read_excel(out_xlsx, engine="openpyxl", dtype=str)
            df_existing = clean_dataframe(df_existing).fillna("").astype(str)
            # マージ：既存分へのカウント＆追記＋履歴つき新規分の追加
            added_records = merge_person_records(df_person2, df_existing, today_str, logger)
            # df_existing＋added_recordsで合成
            df_person2 = pd.concat([df_existing, added_records], ignore_index=True)
        except Exception as e:
            if logger:
                logger.error(f"既存Excelの読み込みに失敗しました: {e}")

    # 必要カラム埋め
    col_order = [
        "管理番号", "氏名", "メールアドレス", "電話番号", "郵便番号",
        "登録住所", "本人確認登録郵便番号", "本人確認登録時住所",
        "請求公演名", "件数", "備考", "履歴"
    ]
    for col in col_order:
        if col not in df_person2.columns:
            df_person2[col] = ""
    df_person2 = df_person2[col_order]
    df_person2 = clean_dataframe(df_person2).astype(str)

    # 一時ファイル経由で書き込み
    tmp_xlsx = out_xlsx.with_name(out_xlsx.stem + "_tmp.xlsx")
    try:
        out_xlsx.parent.mkdir(parents=True, exist_ok=True)
        df_person2.to_excel(tmp_xlsx, index=False)
        tmp_xlsx.replace(out_xlsx)
        style_excel(out_xlsx, CFG["excel"]["font_name"])
    except PermissionError:
        if logger:
            logger.error("CustList.xlsx を開いているため書き込めません。閉じてから再実行してください。")
        return

    # 重複分退避
    # if not removed_records.empty:
    #     export_duplicate_records(removed_records, base_dir=out_xlsx.parent, today_str=ts)

    # ===== CSV出力用 DataFrame作成 =====
    if "管理番号" not in df_new.columns:
        df_new["管理番号"] = ""
    person_to_no = dict(zip(
        zip(df_person2["氏名"], df_person2["メールアドレス"], df_person2["請求公演名"]),
        df_person2["管理番号"]
    ))
    df_new["管理番号"] = df_new.apply(
        lambda row: person_to_no.get((row["氏名"], row["メールアドレス"], row["請求公演名"]), ""),
        axis=1
    )
    # 不要なカラム除去
    csv_col_order = [
        "管理番号", "氏名", "メールアドレス", "電話番号", "郵便番号",
        "登録住所", "本人確認登録郵便番号", "本人確認登録時住所",
        "請求公演名", "備考", "閲覧用URL"
    ]
    for col in csv_col_order:
        if col not in df_new.columns:
            df_new[col] = ""
    csv_df = df_new[csv_col_order]

    # ファイル名例：CustList_2025-06-04_2020.csv
    csv_name = CFG["paths"]["csv_pattern"].replace(
        "{yyyymmdd}", dt.datetime.today().strftime("%Y-%m-%d_%H%M")
    )
    Path(csv_name).parent.mkdir(parents=True, exist_ok=True)
    csv_df = clean_dataframe(csv_df).astype(str)
    tmp_csv_name = Path(csv_name).with_name(Path(csv_name).stem + "_tmp.csv")
    try:
        csv_df.to_csv(tmp_csv_name, index=False, encoding=CFG["csv"]["encoding"])
        tmp_csv_name.replace(csv_name)
    except Exception as e:
        if logger:
            logger.error(f"CSV一時ファイルの書き出しに失敗しました。元ファイルは壊れていません。: {e}")
        return

    if logger:
        logger.info(f"追記完了：{len(df_person2)} 人 / {len(df_new)} URL")
        logger.info(f"Excel 保存: {out_xlsx}")
        logger.info(f"CSV 出力 : {csv_name}")

def load_person_map(logger=None) -> dict[str, str]:
    p = Path(CFG["paths"]["output_excel"])
    if not p.exists():
        return {}
    df = pd.read_excel(p, engine="openpyxl", dtype=str).fillna("").astype(str)
    df = clean_dataframe(df)
    df["key"] = df["氏名"].astype(str) + "|" + df["メールアドレス"].astype(str) + "|" + df["請求公演名"].astype(str)
    return dict(zip(df["key"], df["管理番号"].astype(str)))
